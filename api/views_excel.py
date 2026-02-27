import openpyxl
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl.utils import get_column_letter
from .models import Asset, Employee, Assignment
import uuid

def export_assets_excel(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'GET only'}, status=405)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ativos"
    
    headers = ["ID", "Nome do Equipamento", "Categoria", "Marca", "Condição", "Data de Entrada", "Processador", "Memória RAM", "Armazenamento", "IMEI", "Linha Telefônica", "Observações", "Status", "Funcionário Atual", "Matrícula", "Departamento", "Data de Empréstimo", "Última Devolução", "Termo Assinado"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    assets = Asset.objects.all().prefetch_related('assignments__employee')
    
    for row_num, asset in enumerate(assets, 2):
        active_assignment = asset.assignments.filter(status='active').first()
        last_returned = asset.assignments.filter(status='returned').order_by('-return_date').first()
        
        status = "Em Uso" if active_assignment else "Disponível"
        emp_name = active_assignment.employee.name if active_assignment else "-"
        emp_mat = active_assignment.employee.matricula if active_assignment and active_assignment.employee.matricula else "-"
        emp_dep = active_assignment.employee.department if active_assignment and active_assignment.employee.department else "-"
        date_str = active_assignment.date.strftime("%d/%m/%Y %H:%M") if active_assignment else "-"
        
        last_returned_str = last_returned.return_date.strftime("%d/%m/%Y %H:%M") if last_returned and last_returned.return_date else "-"
        signed = "Sim" if active_assignment and active_assignment.signed_term else ("-" if not active_assignment else "Não")

        entry_date_str = asset.entry_date.strftime("%d/%m/%Y") if asset.entry_date else "-"

        ws.cell(row=row_num, column=1, value=asset.id)
        ws.cell(row=row_num, column=2, value=asset.name)
        ws.cell(row=row_num, column=3, value=asset.category)
        ws.cell(row=row_num, column=4, value=asset.brand or "-")
        ws.cell(row=row_num, column=5, value=asset.condition)
        ws.cell(row=row_num, column=6, value=entry_date_str)
        ws.cell(row=row_num, column=7, value=asset.processor or "-")
        ws.cell(row=row_num, column=8, value=asset.ram_memory or "-")
        ws.cell(row=row_num, column=9, value=asset.storage or "-")
        ws.cell(row=row_num, column=10, value=asset.imei or "-")
        ws.cell(row=row_num, column=11, value=asset.phone_number or "-")
        ws.cell(row=row_num, column=12, value=asset.observation or "-")
        ws.cell(row=row_num, column=13, value=status)
        ws.cell(row=row_num, column=14, value=emp_name)
        ws.cell(row=row_num, column=15, value=emp_mat)
        ws.cell(row=row_num, column=16, value=emp_dep)
        ws.cell(row=row_num, column=17, value=date_str)
        ws.cell(row=row_num, column=18, value=last_returned_str)
        ws.cell(row=row_num, column=19, value=signed)
        
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ativos_cooperativa.xlsx"'
    wb.save(response)
    return response

@csrf_exempt
def import_assets_excel(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
        
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
        
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] and not row[1]: # Ignora linhas totalmente vazias
                continue
                
            # No nosso caso simples, se vier Nome e Categoria já salva
            # row[0] poderia ser ID mas vamos ignorar ou recriar.
            # Assumimos Formato: Coluna B=Nome, Coluna C=Categoria
            
            # Se for formato antigo ou sem id:
            name = row[1] if len(row) > 1 and row[1] else str(row[0] or "Ativo Sem Nome")
            category = row[2] if len(row) > 2 and row[2] else "Outros"
            
            Asset.objects.create(name=name, category=category)
            count += 1
            
        return JsonResponse({'success': True, 'imported': count})
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)
